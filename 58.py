import time
import requests
import openpyxl
from openpyxl.styles import Alignment

from lxml import etree

HEADERS = {'user-agent' :'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.121 Safari/537.36'}

url = 'https://zz.58.com/ershoufang/pn%d/'




def sc_page():
    for i in range(1,31):
        new = format(url%i)
        page_text = requests.get(new, headers=HEADERS).text
        tree = etree.HTML(page_text)
        li_list = tree.xpath('//ul[@class="house-list-wrap"]//li')
        for li in li_list:
            title = li.xpath('./div[2]/h2/a/text()')[0] # 名称
            layout = li.xpath('./div[2]/p[1]/span[1]/text()')[0] # 房屋布局
            area = li.xpath('./div[2]/p[1]/span[2]/text()')[0]  # 面积
            position_list = li.xpath('./div[2]/p[2]/span[1]//text()')
            position = ''.join(''.join(position_list).split())  # 位置
            price = ''.join(li.xpath('./div[3]/p[1]//text()'))   #总价
            unit = li.xpath('./div[3]/p[2]/text()')[0]  #单价
            intermediary = li.xpath('./div[2]/div/span/text()')[0] # 中介公司
            #agent = li.xpath('./div[2]/div/a/span/text()')[0] or str(li.xpath('./div[2]/div/a/span/text()')) # 经纪人
            ws.append([title, layout, area, position, price, unit, intermediary])
        print(new + '下载完成')

if __name__ == "__main__":
    

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.merge_cells('A1:G1')
    ws.cell(1,1).value = '郑州市58同城二手房'
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.append(['房屋名称', '房屋布局', '房屋面积', '位置', '总价', '单价', '中介公司'])

    sc_page()

    wb.save('1.xlsx')

